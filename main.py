# Link to original code: https://automatetheboringstuff.com/2e/chapter13/

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font

 # Load produceSales workbook, and freeze row 1
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

sheet.freeze_panes = 'A2'

# Update prices in the produceSales workbook
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        sheet.cell(row=rowNum, column=2).font = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

# Create a pivot table
df_produce = pd.read_excel("produceSales.xlsx")
df_produce
produce_pivot = df_produce.pivot_table(values="TOTAL", index="PRODUCE", aggfunc="sum")

# Save it to a new excel workbook
produce_pivot.to_excel('pivot_table.xlsx')

# Create a bar plot from the pivot table
produce_plot = produce_pivot.plot(kind='bar', xlabel='', ylabel='', title='', figsize=(8,4))
plt.savefig('produce.png')
plt.show()